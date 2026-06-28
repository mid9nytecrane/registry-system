from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse

from .models import PartyMember, PollingStation,ElectoralArea
from .forms import PartyMemberForm,PollingStationForm

from core.resource import PartyMemberResource

import io 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────
#  Auth
# ─────────────────────────────────────────

def is_admin(user):
    return user.is_staff

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect(request.GET.get('next', 'home'))
        messages.error(request, 'Invalid username or password.')

    # Minimal form-like object so the template can access {{ form.username.value }}
    class _FakeForm:
        class username:
            @staticmethod
            def value():
                return request.POST.get('username', '')
        non_field_errors = []

    return render(request, 'core/login.html', {'form': _FakeForm()})


def logout_view(request):
    logout(request)
    return redirect('login')


# ─────────────────────────────────────────
#  Dashboard
# ─────────────────────────────────────────

@login_required
def home_view(request):
    now = timezone.now()
    recent_members = PartyMember.objects.select_related('polling_station').order_by('-date_registered')[:5]
    total_members = PartyMember.objects.count()
    total_stations = PollingStation.objects.count()
    new_this_month = PartyMember.objects.filter(
        date_registered__year=now.year,
        date_registered__month=now.month,
    ).count()
    active_stations = PollingStation.objects.filter(members__isnull=False).distinct().count()

    return render(request, 'core/home.html', {
        'recent_members': recent_members,
        'total_members': total_members,
        'total_stations': total_stations,
        'new_this_month': new_this_month,
        'active_stations': active_stations,
    })


# ─────────────────────────────────────────
#  Members
# ─────────────────────────────────────────

@login_required
def member_list_view(request):
    qs = PartyMember.objects.select_related('polling_station').order_by('-date_registered')

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(member_id__icontains=q) |
            Q(phone__icontains=q)
        )

    station_id = request.GET.get('station', '')
    if station_id:
        qs = qs.filter(polling_station_id=station_id)

    return render(request, 'core/member_list.html', {
        'members': qs,
        'stations': PollingStation.objects.all(),
        'q': q,
        'selected_station': station_id,
    })


@login_required
def member_detail_view(request, pk):
    member = get_object_or_404(PartyMember, pk=pk)
    return render(request, 'core/member_detail.html', {'member': member})


@login_required
def member_create_view(request):
    # Pre-select station from query param if provided
    initial = {}
    preselected = request.GET.get('station', '')
    if preselected:
        initial['polling_station'] = preselected

    if request.method == 'POST':
        form = PartyMemberForm(request.POST)
        if form.is_valid():
            member = form.save(commit=False)
            member.registered_by = request.user
            member.save()
            messages.success(request, f'Member {member.first_name} {member.last_name} registered successfully.')
            return redirect('member_detail', pk=member.pk)
    else:
        form = PartyMemberForm(initial=initial)

    return render(request, 'core/member_form.html', {'form': form})


@login_required
def member_edit_view(request, pk):
    member = get_object_or_404(PartyMember, pk=pk)

    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to edit member records.')
        return redirect('member_detail', pk=pk)

    if request.method == 'POST':
        form = PartyMemberForm(request.POST, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, 'Member record updated successfully.')
            return redirect('member_detail', pk=member.pk)
    else:
        form = PartyMemberForm(instance=member)

    return render(request, 'core/member_form.html', {
        'form': form,
        'member': member,
    })


# ─────────────────────────────────────────
#  Polling Stations
# ─────────────────────────────────────────

@login_required
def station_list_view(request):

    query = request.GET.get("q", "").strip()
    stations = PollingStation.objects.all()
    if query:
        stations = stations.filter(
            Q(name__icontains=query) | Q(location__icontains=query)
        )
   
    return render(request, 'core/station_list.html', {'stations': stations, 'query':query})


@login_required
def station_detail_view(request, pk):
    station = get_object_or_404(PollingStation, pk=pk)
    members = station.members.order_by('last_name', 'first_name')
    
    return render(request, 'core/station_detail.html', {
        'station': station,
        'members': members,
    })

@login_required
def station_create_view(request):
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Only admins can add polling stations.')
        return redirect('station_list')

    initial = {}
    preselected = request.GET.get('electoral_area', '')
    if preselected:
        initial['electoral_area'] = preselected

    if request.method == 'POST':
        form = PollingStationForm(request.POST)
        if form.is_valid():
            station = form.save()
            messages.success(request, f'Polling station "{station.name}" added successfully.')
            return redirect('station_detail', pk=station.pk)
    else:
        form = PollingStationForm(initial=initial)

    # Pass all electoral area data as JSON for JS auto-population
    import json
    areas_data = {
        str(area.pk): {
            'location':     area.location or '',
            'constituency': area.constituency or '',
            'description':  area.description or '',
        }
        for area in ElectoralArea.objects.all()
    }

    return render(request, 'core/station_form.html', {
        'form': form,
        'areas_json': json.dumps(areas_data),
    })


@login_required
def station_edit_view(request, pk):
    station = get_object_or_404(PollingStation, pk=pk)

    if not request.user.is_staff:
        messages.error(request, 'Only administrators can edit polling stations.')
        return redirect('station_detail', pk=pk)

    if request.method == 'POST':
        form = PollingStationForm(request.POST, instance=station)
        if form.is_valid():
            form.save()
            messages.success(request, f'Station "{station.name}" updated successfully.')
            return redirect('station_detail', pk=station.pk)
    else:
        form = PollingStationForm(instance=station)

    import json
    areas_data = {
        str(area.pk): {
            'location':     area.location or '',
            'constituency': area.constituency or '',
            'description':  area.description or '',
        }
        for area in ElectoralArea.objects.all()
    }

    return render(request, 'core/station_form.html', {
        'form':       form,
        'station':    station,
        'areas_json': json.dumps(areas_data),
    })


# ─────────────────────────────────────────
#  Electoral Areas
# ─────────────────────────────────────────
@login_required
def electoral_area_list_view(request):
    query = request.GET.get('q', '').strip()
    electoral_areas = ElectoralArea.objects.all()
    if query:
        electoral_areas = electoral_areas.filter(name__icontains=query)
    
    context = {
        'electoral_areas': electoral_areas
    }
    return render(request, 'core/area_list.html', context)


@login_required
def electoral_area_create_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Only master admin can create electoral area." )
        return redirect('area_list')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, "Electoral Area name is required")
        else:
            area = ElectoralArea.objects.create(
                name=name,
                location=request.POST.get('location', '').strip(),
                constituency=request.POST.get('constituency', '').strip(),
                description=request.POST.get('description', '').strip(),
            )

            messages.success(request, f"Electoral Area '{area.name}' is created succesfully.")
            return redirect('area_detail', pk=area.pk)
    return render(request, 'core/area_form.html')



@login_required
def electoral_area_details_view(request, pk):
    electoral_area = get_object_or_404(ElectoralArea, pk=pk)
    stations = electoral_area.stations.order_by('name')
    print(f"polling stations under {electoral_area}: {stations}") 

    context = {
        'stations':stations
    }

    return render(request, 'core/area_detail.html', context)



@login_required  # BUG 1: missing @ decorator
def electoral_area_edit_view(request, pk):
    area = get_object_or_404(ElectoralArea, pk=pk)

    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can edit electoral areas.')
        return redirect('area_list')  # BUG 2: area_list takes no pk argument

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Electoral Area name is required.')
        else:
            area.name = name
            area.location = request.POST.get('location', '').strip()
            area.constituency = request.POST.get('constituency', '').strip()
            area.description = request.POST.get('description', '').strip()
            area.save()
            messages.success(request, f'Electoral Area "{area.name}" updated successfully.')
            return redirect('area_list')

    # BUG 3: render was missing the context — area was never passed to the template
    # so {{ area.name }}, {{ area.location }} etc. all rendered empty
    return render(request, 'core/area_form.html', {'area': area})



# ─────────────────────────────────────────
#  Admin Dashboard
# ─────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    """Custom admin panel — staff only. Superusers access Django /admin/."""
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access the admin panel.')
        return redirect('home')

    now = timezone.now()

    # Aggregate stats
    total_members    = PartyMember.objects.count()
    total_stations   = PollingStation.objects.count()
    new_this_month   = PartyMember.objects.filter(
        date_registered__year=now.year,
        date_registered__month=now.month,
    ).count()
    male_count       = PartyMember.objects.filter(gender='M').count()
    female_count     = PartyMember.objects.filter(gender='F').count()

    # All users except superusers (they manage themselves via Django admin)
    from django.contrib.auth.models import User
    users = User.objects.filter(is_superuser=False).order_by('-date_joined')

    # Recent registrations
    recent_members = (
        PartyMember.objects
        .select_related('polling_station', 'registered_by')
        .order_by('-date_registered')[:10]
    )

    # Stations with member counts
    stations = PollingStation.objects.all()

    return render(request, 'core/admin_dashboard.html', {
        'total_members':  total_members,
        'total_stations': total_stations,
        'new_this_month': new_this_month,
        'male_count':     male_count,
        'female_count':   female_count,
        'users':          users,
        'recent_members': recent_members,
        'stations':       stations,
    })


@login_required
def admin_user_toggle_view(request, user_id):
    """Toggle a user's is_staff status — superuser only."""
    if not request.user.is_superuser:
        messages.error(request, 'Only the master administrator can change user roles.')
        return redirect('admin_dashboard')

    from django.contrib.auth.models import User
    user = get_object_or_404(User, pk=user_id)
    if user == request.user:
        messages.error(request, 'You cannot change your own role.')
        return redirect('admin_dashboard')

    user.is_staff = not user.is_staff
    user.save()
    role = 'Administrator' if user.is_staff else 'Member'
    messages.success(request, f'{user.get_full_name() or user.username} is now a {role}.')
    return redirect('admin_dashboard')


@login_required
def admin_user_delete_view(request, user_id):
    """Delete a user account — superuser only."""
    if not request.user.is_superuser:
        messages.error(request, 'Only the master administrator can delete users.')
        return redirect('admin_dashboard')

    from django.contrib.auth.models import User
    user = get_object_or_404(User, pk=user_id)
    if user == request.user:
        messages.error(request, 'You cannot delete your own account.')
        return redirect('admin_dashboard')

    name = user.get_full_name() or user.username
    user.delete()
    messages.success(request, f'User "{name}" has been deleted.')
    return redirect('admin_dashboard')


# helpers for exporting data
def _build_membership_row(station):
    members_per_station = PartyMember.objects.filter(polling_station=station)
    headers = ['#', 'First Name', 'Middle Name', 'Last Name', 'Party ID', 'Voters ID', 'Phone']

    rows = []

    for i, member in enumerate(members_per_station, start=1):
        rows.append([
            i,
            member.first_name,
            member.middle_name or '-',
            member.last_name,
            member.member_id,
            member.voters_id,
            member.phone
        ])

    return headers,rows


#exporting data
@login_required
@user_passes_test(is_admin)
def export_data(request,pk):
    station = get_object_or_404(PollingStation, pk=pk)
    headers, rows = _build_membership_row(station)
    queryset = PartyMember.objects.filter(polling_station=station)
    dataset = PartyMemberResource().export(queryset=queryset)

    #---------------------------------------------
    # Styling exported Excel Sheet
    #----------------------------------------------

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{station.name} Members"

    # ---- color palette -------------------------
    green       = "198A19"
    white       = "FFFFFF"

    #3. Style definitions
    header_font    = Font(name='Arial',bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill(fill_type="solid", fgColor="2d5a27")
    header_align   = Alignment(horizontal="center", vertical="center")
    data_align     = Alignment(horizontal="left", vertical="center")
    stripe_fill    = PatternFill(fill_type="solid", fgColor="EAF4E8")
    thin           = Side(style="thin", color="CCCCCC")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----------- title block -------------
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"{station.name} Membership Database"
    title_cell.font = Font(name='Times New Roman', bold=True, size=16, color=white)
    title_cell.fill = PatternFill("solid", fgColor=green)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.append([]) #blank space

    # Define columns: (header label)
    # columns = [
    #     ("First Name", lambda m: m.first_name),
    #     ("Middle Name", lambda m: m.middle_name),
    #     ("Last Name", lambda m: m.last_name),
    #     ("Party ID", lambda m: m.member_id),
    #     ("Voters ID", lambda m: m.voters_id),
    #     ("Phone", lambda m: m.phone)
    # ]

    # ----------- header row (row 3) -------------------
    ws.append(headers)
    header_row = ws[3]

    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill 
        cell.alignment = header_align
    ws.row_dimensions[3].height=22
    

    #----------- data row (starting from row 4) ----------------
    for row_idx, row_data in enumerate(rows, start=4):
        ws.append(row_data)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.alignment = data_align
            cell.border = border 
            if row_idx % 2 == 0:
                cell.fill = stripe_fill

     # ----------- auto-fit column widths ---------
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
            for cell in row:
                max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = max_length + 4

    # freeze the header row i.e like static navbar type sh!!t 
    ws.freeze_panes = "A4"

    #save and send 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    
    response = HttpResponse(
        buffer.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f"attachment; filename='{station.name}_members.xlsx"
    return response
